import json
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import re

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Rural Connectivity Budget Tool", page_icon="🧮", layout="wide")
st.logo("./assets/fikalogo.png")
st.sidebar.header("Rural Connectivity Budget Tool")

WORKBOOK_CANDIDATES = [
    Path("./data/Rural Connectivity Budget Tool.xlsx"),
    Path("./Networked Transport Infrastructure Budget Allocation Tool V2.xlsx"),
]
WORKBOOK_PATH = next((path for path in WORKBOOK_CANDIDATES if path.exists()), WORKBOOK_CANDIDATES[0])
DATA_JSON_PATH = Path("./data/rnti_budget_calculator_data.json")

FALLBACK_SOURCE_ROW_GROUPS = {
    25: [7, 8, 9, 10],
    26: [16, 17, 18, 19, 20, 21],
    27: [26, 27],
    28: [32, 33, 34, 35, 36],
    29: [41, 42],
    30: [47, 48],
    33: [54],
    34: [60, 61],
    35: [66, 67],
    36: [72],
    37: [77, 78],
    38: [83],
    39: [88],
    40: [93],
    42: [99, 100, 101, 102],
    43: [107, 108, 109],
    44: [114, 115],
    45: [120],
    46: [125],
    47: [130, 131],
    48: [136, 137],
    49: [148],
    50: [142, 143],
}


def normalize_cell_ref(ref: str) -> str:
    ref = ref.strip().replace("$", "")
    if "!" in ref:
        ref = ref.split("!", 1)[1]
    return ref


def extract_average_refs(formula: str | None) -> list[str]:
    if not isinstance(formula, str) or "AVERAGE(" not in formula.upper():
        return []

    match = re.search(r"AVERAGE\((.*?)\)", formula, flags=re.IGNORECASE)
    if not match:
        return []

    inside = match.group(1)
    refs = [part.strip() for part in inside.split(",") if part.strip()]
    expanded: list[str] = []

    for ref in refs:
        if ":" in ref:
            start_ref_raw, end_ref_raw = ref.split(":", 1)
            start_ref = normalize_cell_ref(start_ref_raw)
            end_ref = normalize_cell_ref(end_ref_raw)
            start_match = re.match(r"([A-Za-z]+)(\d+)", start_ref)
            end_match = re.match(r"([A-Za-z]+)(\d+)", end_ref)
            if not start_match or not end_match:
                continue

            start_col, start_row = start_match.group(1).upper(), int(start_match.group(2))
            end_col, end_row = end_match.group(1).upper(), int(end_match.group(2))
            if start_col != end_col:
                continue

            low, high = sorted((start_row, end_row))
            expanded.extend([f"{start_col}{row}" for row in range(low, high + 1)])
        else:
            single_ref = normalize_cell_ref(ref)
            if re.match(r"[A-Za-z]+\d+", single_ref):
                expanded.append(single_ref.upper())

    return sorted(set(expanded))


def extract_sheet_cell_row(formula: str | None, col_letter: str) -> int | None:
    if not isinstance(formula, str):
        return None
    match = re.search(rf"!\$?{col_letter}(\d+)", formula, flags=re.IGNORECASE)
    if not match:
        return None
    return int(match.group(1))


def compile_excel_formula(formula: str | None):
    """Compile selected Excel formulas into Python callables.

    Supported patterns:
    - IFERROR(ROUND(AVERAGE(...),0),"")
    - IFERROR(ROUND((A1+B1)/2,0),"")
    """
    if not isinstance(formula, str):
        return None

    avg_refs = extract_average_refs(formula)
    if avg_refs:
        def eval_average(resolve_ref):
            values = [resolve_ref(ref) for ref in avg_refs]
            return calculate_blended_average(pd.Series(values))

        return eval_average

    normalized = formula.replace("$", "")
    mid_match = re.search(
        r"ROUND\(\s*\(?\s*([A-Za-z]+\d+)\s*\+\s*([A-Za-z]+\d+)\s*\)?\s*/\s*2\s*,\s*0\s*\)",
        normalized,
        flags=re.IGNORECASE,
    )
    if mid_match:
        left_ref = normalize_cell_ref(mid_match.group(1)).upper()
        right_ref = normalize_cell_ref(mid_match.group(2)).upper()

        def eval_mid(resolve_ref):
            return calculate_source_mid(resolve_ref(left_ref), resolve_ref(right_ref))

        return eval_mid

    return None


def parse_average_source_rows(formula: str | None, col_letter: str) -> list[int]:
    refs = extract_average_refs(formula)
    rows: list[int] = []
    for ref in refs:
        match = re.match(r"([A-Za-z]+)(\d+)", ref)
        if not match:
            continue
        col, row = match.group(1).upper(), int(match.group(2))
        if col == col_letter.upper():
            rows.append(row)
    return sorted(set(rows))


def derive_source_row_groups(ws_budget_formula, ws_source_formula) -> dict[int, list[int]]:
    row_groups: dict[int, list[int]] = {}

    for row in range(24, 51):
        item_number = ws_budget_formula[f"A{row}"].value
        if item_number is None:
            continue

        min_formula = ws_budget_formula[f"D{row}"].value
        source_rows: list[int] = []

        blended_row = extract_sheet_cell_row(min_formula, "C")
        if blended_row is not None:
            source_rows = parse_average_source_rows(ws_source_formula[f"C{blended_row}"].value, "C")
            if not source_rows:
                source_rows = parse_average_source_rows(ws_source_formula[f"D{blended_row}"].value, "D")

        if not source_rows:
            source_rows = FALLBACK_SOURCE_ROW_GROUPS.get(row, [])

        if source_rows:
            row_groups[row] = source_rows

    return row_groups


def derive_budget_to_source_blended_row(ws_budget_formula) -> dict[int, int]:
    links: dict[int, int] = {}
    for row in range(24, 51):
        item_number = ws_budget_formula[f"A{row}"].value
        if item_number is None:
            continue
        min_formula = ws_budget_formula[f"D{row}"].value
        if not isinstance(min_formula, str):
            continue
        blended_row = extract_sheet_cell_row(min_formula, "C")
        if blended_row is not None:
            links[row] = blended_row
    return links


def source_group_display_order(source_row_groups: dict[int, list[int]]) -> list[int]:
    return sorted(
        source_row_groups.keys(),
        key=lambda budget_row: min(source_row_groups.get(budget_row, [10**9])),
    )


def normalize_json_value(value):
    if value is None:
        return None
    if isinstance(value, (str, int, float, bool)):
        if isinstance(value, float) and pd.isna(value):
            return None
        return value

    try:
        if pd.isna(value):
            return None
    except TypeError:
        pass

    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)

    return str(value)


def dataframe_to_json_records(df: pd.DataFrame) -> list[dict]:
    records = df.to_dict(orient="records")
    return [
        {k: normalize_json_value(v) for k, v in record.items()}
        for record in records
    ]


def save_budget_data_json(
    json_path: str,
    intro_title: str,
    intro_text: str,
    disclaimer: str,
    defaults: dict,
    components_df: pd.DataFrame,
    source_df: pd.DataFrame,
    source_row_groups: dict[int, list[int]],
    budget_to_source_blended_row: dict[int, int],
    section_1_formulas: dict[str, str | None],
    section_3_formulas: dict[str, str | None],
):
    json_file = Path(json_path)
    json_file.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "intro_title": intro_title,
        "intro_text": intro_text,
        "disclaimer": disclaimer,
        "defaults": defaults,
        "components": dataframe_to_json_records(components_df),
        "source": dataframe_to_json_records(source_df),
        "source_row_groups": {str(k): v for k, v in source_row_groups.items()},
        "budget_to_source_blended_row": {
            str(k): v for k, v in budget_to_source_blended_row.items()
        },
        "section_1_formulas": section_1_formulas,
        "section_3_formulas": section_3_formulas,
    }
    json_file.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


@st.cache_data
def load_budget_calculator_data_from_workbook(
    file_path: str,
    workbook_mtime: float,
    json_path: str,
):
    wb = load_workbook(file_path, data_only=True)
    wb_formula = load_workbook(file_path, data_only=False)
    ws = wb["Budget Calculator"]
    ws_source = wb["Source Data - In Progress"]
    ws_formula = wb_formula["Budget Calculator"]
    ws_source_formula = wb_formula["Source Data - In Progress"]
    source_row_groups = derive_source_row_groups(ws_formula, ws_source_formula)
    budget_to_source_blended_row = derive_budget_to_source_blended_row(ws_formula)

    def source_url_from_row(source_row: int):
        cell = ws_source_formula[f"F{source_row}"]
        if cell.hyperlink is None:
            return None
        if cell.hyperlink.target:
            return str(cell.hyperlink.target)
        if cell.hyperlink.location:
            return f"#{cell.hyperlink.location}"
        return None

    def blended_costs_from_source_rows(source_rows: list[int], blended_row: int | None):
        def resolve_source_ref(ref: str):
            normalized_ref = normalize_cell_ref(ref)
            return ws_source[normalized_ref].value

        if blended_row is not None:
            min_formula_fn = compile_excel_formula(ws_source_formula[f"C{blended_row}"].value)
            max_formula_fn = compile_excel_formula(ws_source_formula[f"D{blended_row}"].value)
            mid_formula_fn = compile_excel_formula(ws_source_formula[f"E{blended_row}"].value)

            if min_formula_fn and max_formula_fn:
                min_cost = min_formula_fn(resolve_source_ref)
                max_cost = max_formula_fn(resolve_source_ref)
                if mid_formula_fn:
                    mid_cost = mid_formula_fn(resolve_source_ref)
                else:
                    mid_cost = calculate_source_mid(min_cost, max_cost)
                return min_cost, max_cost, mid_cost

            # Handles direct links like =IF('Source Data - In Progress'!C15="","",'Source Data - In Progress'!C15)
            # where C/D/E are already the blended values in Source Data.
            direct_min = ws_source[f"C{blended_row}"].value
            direct_max = ws_source[f"D{blended_row}"].value
            direct_mid = ws_source[f"E{blended_row}"].value

            if direct_min is not None and direct_max is not None:
                return direct_min, direct_max, calculate_source_mid(direct_min, direct_max)
            if direct_min is not None:
                return direct_min, direct_min, direct_mid if direct_mid is not None else direct_min

        low_values = [ws_source[f"C{r}"].value for r in source_rows]
        high_values = [ws_source[f"D{r}"].value for r in source_rows]

        min_cost = calculate_blended_average(pd.Series(low_values))
        max_cost = calculate_blended_average(pd.Series(high_values))
        mid_cost = calculate_source_mid(min_cost, max_cost)
        return min_cost, max_cost, mid_cost

    intro_title = ws["A1"].value or ""
    intro_text = ws["A2"].value or ""
    disclaimer = ws["A4"].value or ""

    defaults = {
        "project_name": ws["C9"].value or "",
        "country_region": ws["C10"].value or "",
        "total_budget_usd": float(ws["C12"].value or 0),
        "allocation_pct": float(ws["C13"].value or 0),
        "maintenance_pct": float(ws["C16"].value or 0),
    }

    section_1_formulas = {
        "rnti_budget_available": ws_formula["C14"].value,
        "maintenance_reserve_amount": ws_formula["C17"].value,
        "capital_budget_for_infrastructure": ws_formula["C18"].value,
    }

    section_3_formulas = {
        "total_infrastructure_cost": ws_formula["C55"].value,
        "capital_budget_available": ws_formula["C56"].value,
        "remaining_capital_budget": ws_formula["C57"].value,
        "pct_capital_budget_used": ws_formula["C58"].value,
        "maintenance_reserve": ws_formula["C59"].value,
        "total_estimated_rnti_investment": ws_formula["C60"].value,
        "pct_total_corridor_project": ws_formula["C61"].value,
    }

    section_name = ""
    items = []
    for row in range(24, 51):
        item_number = ws[f"A{row}"].value
        infra_name = ws[f"B{row}"].value

        if item_number is None and infra_name:
            section_name = str(infra_name)
            continue

        if item_number is None and not infra_name:
            continue

        source_rows = source_row_groups.get(row, [])
        blended_row = extract_sheet_cell_row(ws_formula[f"D{row}"].value, "C")
        min_cost, max_cost, mid_cost = blended_costs_from_source_rows(source_rows, blended_row)

        items.append(
            {
                "Budget Row": row,
                "#": int(item_number),
                "Section": section_name,
                "Infrastructure Type": str(infra_name),
                "Unit": ws[f"C{row}"].value,
                "Min Unit Cost (USD)": min_cost,
                "Max Unit Cost (USD)": max_cost,
                "Mid Unit Cost (USD)": mid_cost,
                "Qty": float(ws[f"G{row}"].value or 0),
            }
        )

    components_df = pd.DataFrame(items)

    def section_title_for_source_row(row_num: int) -> str:
        if row_num < 51:
            return "ORDER 3 TERTIARY LINKS"
        if row_num < 96:
            return "SUB-TERTIARY LINKS"
        return "ANCILLARY & SAFETY INFRASTRUCTURE"

    source_records = []
    for budget_row, source_rows in source_row_groups.items():
        infrastructure_type = ws[f"B{budget_row}"].value
        first_source_row = min(source_rows) if source_rows else None
        item_title = ws_source[f"A{first_source_row - 2}"].value if first_source_row else None
        section_title = section_title_for_source_row(first_source_row) if first_source_row else ""
        for source_row in source_rows:
            source_url = source_url_from_row(source_row)
            source_records.append(
                {
                    "Budget Row": budget_row,
                    "Source Section": section_title,
                    "Source Item Title": item_title,
                    "Infrastructure Type": infrastructure_type,
                    "Source Row": source_row,
                    "Source #": ws_source[f"A{source_row}"].value,
                    "Geography / Context": ws_source[f"B{source_row}"].value,
                    "Low Est. (USD)": ws_source[f"C{source_row}"].value,
                    "High Est. (USD)": ws_source[f"D{source_row}"].value,
                    "Source Mid (USD)": ws_source[f"E{source_row}"].value,
                    "Source Name": source_url or ws_source[f"F{source_row}"].value,
                    "Full Citation": ws_source[f"G{source_row}"].value,
                    "Year": ws_source[f"H{source_row}"].value,
                }
            )

    source_df = pd.DataFrame(source_records)
    save_budget_data_json(
        json_path,
        intro_title,
        intro_text,
        disclaimer,
        defaults,
        components_df,
        source_df,
        source_row_groups,
        budget_to_source_blended_row,
        section_1_formulas,
        section_3_formulas,
    )

    return (
        intro_title,
        intro_text,
        disclaimer,
        defaults,
        components_df,
        source_df,
        source_row_groups,
        budget_to_source_blended_row,
        section_1_formulas,
        section_3_formulas,
    )


@st.cache_data
def load_budget_calculator_data_from_json(json_path: str, json_mtime: float):
    payload = json.loads(Path(json_path).read_text(encoding="utf-8"))
    intro_title = payload.get("intro_title", "")
    intro_text = payload.get("intro_text", "")
    disclaimer = payload.get("disclaimer", "")
    defaults = payload.get("defaults", {})
    components_df = pd.DataFrame(payload.get("components", []))
    source_df = pd.DataFrame(payload.get("source", []))
    source_row_groups = {
        int(k): v for k, v in payload.get("source_row_groups", {}).items()
    }
    budget_to_source_blended_row = {
        int(k): int(v)
        for k, v in payload.get("budget_to_source_blended_row", {}).items()
    }
    section_1_formulas = payload.get("section_1_formulas", {})
    section_3_formulas = payload.get("section_3_formulas", {})

    return (
        intro_title,
        intro_text,
        disclaimer,
        defaults,
        components_df,
        source_df,
        source_row_groups,
        budget_to_source_blended_row,
        section_1_formulas,
        section_3_formulas,
    )


def calc_range_text(min_unit_cost: float | None, max_unit_cost: float | None, qty: float):
    if qty == 0:
        return "-"
    if pd.isna(min_unit_cost) or pd.isna(max_unit_cost):
        return "No verifiable source yet"
    return f"${min_unit_cost * qty:,.0f} - ${max_unit_cost * qty:,.0f}"


def to_numeric_series(values: pd.Series) -> pd.Series:
    return pd.to_numeric(
        values.astype(str).str.replace(r"[\$,]", "", regex=True).str.strip(),
        errors="coerce",
    )


def excel_round(value: float, digits: int = 0) -> float:
    """Mimic Excel ROUND behavior (half away from zero)."""
    quantize_exp = Decimal("1").scaleb(-digits)
    return float(Decimal(str(value)).quantize(quantize_exp, rounding=ROUND_HALF_UP))


def build_average_formula(col_letter: str, rows: list[int]) -> str:
    def build_range_ref() -> str:
        if len(rows) == 1:
            return f"{col_letter}{rows[0]}"
        if rows == list(range(min(rows), max(rows) + 1)):
            return f"{col_letter}{min(rows)}:{col_letter}{max(rows)}"
        return ",".join([f"{col_letter}{r}" for r in rows])

    if not rows:
        return ""
    range_ref = build_range_ref()
    return f'=IFERROR(ROUND(AVERAGE({range_ref}),0),"")'


def build_mid_formula(rows: list[int]) -> str:
    if not rows:
        return ""
    c_formula = build_average_formula("C", rows).lstrip("=")
    d_formula = build_average_formula("D", rows).lstrip("=")
    return f'=IFERROR(ROUND((({c_formula})+({d_formula}))/2,0),"")'


def build_blended_summary_table(source_data: pd.DataFrame, source_row_groups: dict[int, list[int]]) -> pd.DataFrame:
    records = []
    for budget_row, source_rows in source_row_groups.items():
        subset = source_data[source_data["Budget Row"] == budget_row].copy()
        infra_name = (
            subset["Infrastructure Type"].iloc[0]
            if not subset.empty
            else None
        )

        lows = subset["Low Est. (USD)"] if not subset.empty else pd.Series(dtype=float)
        highs = subset["High Est. (USD)"] if not subset.empty else pd.Series(dtype=float)

        blended_min = calculate_blended_average(lows)
        blended_max = calculate_blended_average(highs)
        blended_mid = calculate_source_mid(blended_min, blended_max)

        records.append(
            {
                "Budget Row": budget_row,
                "Infrastructure Type": infra_name,
                "Blended Min (USD)": blended_min,
                "Blended Max (USD)": blended_max,
                "Blended Mid (USD)": blended_mid,
                "Blended Min Formula": build_average_formula("C", source_rows),
                "Blended Max Formula": build_average_formula("D", source_rows),
                "Blended Mid Formula": build_mid_formula(source_rows),
            }
        )
    return pd.DataFrame(records)


def calculate_source_mid(low_value, high_value):
    low_num = to_numeric_series(pd.Series([low_value])).iloc[0]
    high_num = to_numeric_series(pd.Series([high_value])).iloc[0]
    if pd.isna(low_num) or pd.isna(high_num):
        return pd.NA
    return excel_round((low_num + high_num) / 2, 0)


def calculate_blended_average(values: pd.Series):
    numeric_values = to_numeric_series(values).dropna()
    if numeric_values.empty:
        return pd.NA
    return excel_round(numeric_values.mean(), 0)


def parse_sum_rows(formula: str | None, col_letter: str) -> list[int]:
    if not isinstance(formula, str) or "SUM(" not in formula.upper():
        return []

    match = re.search(r"SUM\((.*?)\)", formula, flags=re.IGNORECASE)
    if not match:
        return []

    refs = [part.strip() for part in match.group(1).split(",") if part.strip()]
    rows: list[int] = []
    for ref in refs:
        if ":" in ref:
            start_ref_raw, end_ref_raw = ref.split(":", 1)
            start_ref = normalize_cell_ref(start_ref_raw)
            end_ref = normalize_cell_ref(end_ref_raw)
            start_match = re.match(r"([A-Za-z]+)(\d+)", start_ref)
            end_match = re.match(r"([A-Za-z]+)(\d+)", end_ref)
            if not start_match or not end_match:
                continue
            start_col, start_row = start_match.group(1).upper(), int(start_match.group(2))
            end_col, end_row = end_match.group(1).upper(), int(end_match.group(2))
            if start_col != col_letter.upper() or end_col != col_letter.upper():
                continue
            low, high = sorted((start_row, end_row))
            rows.extend(range(low, high + 1))
        else:
            single_ref = normalize_cell_ref(ref)
            single_match = re.match(r"([A-Za-z]+)(\d+)", single_ref)
            if not single_match:
                continue
            if single_match.group(1).upper() == col_letter.upper():
                rows.append(int(single_match.group(2)))

    return sorted(set(rows))


def calculate_total_infrastructure_cost_from_formula(
    formula: str | None,
    budget_row_totals: dict[int, float],
) -> float:
    sum_rows = parse_sum_rows(formula, "H")
    if not sum_rows:
        values = list(budget_row_totals.values())
        return float(pd.Series(values).fillna(0).sum())

    matched_values = [budget_row_totals[row] for row in sum_rows if row in budget_row_totals]
    if matched_values:
        return float(pd.Series(matched_values).fillna(0).sum())

    # Fallback if row keys are unavailable from the editor output.
    values = list(budget_row_totals.values())
    return float(pd.Series(values).fillna(0).sum())


def unwrap_iferror_expression(expression: str) -> str:
    expr = expression.strip()
    if not expr.upper().startswith("IFERROR(") or not expr.endswith(")"):
        return expr

    inner = expr[len("IFERROR("):-1]
    depth = 0
    for i, ch in enumerate(inner):
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        elif ch == "," and depth == 0:
            return inner[:i].strip()

    return inner.strip()


def evaluate_excel_numeric_formula(
    formula: str | None,
    cell_values: dict[str, float],
    fallback: float,
) -> float:
    if not isinstance(formula, str) or not formula.strip():
        return float(fallback)

    expression = formula.strip()
    if expression.startswith("="):
        expression = expression[1:]

    expression = unwrap_iferror_expression(expression).replace("$", "")

    def replace_ref(match: re.Match) -> str:
        ref = match.group(1).upper()
        value = cell_values.get(ref, 0.0)
        return str(float(value or 0.0))

    expression = re.sub(r"\b([A-Za-z]+\d+)\b", replace_ref, expression)
    expression = expression.replace("^", "**")

    if not re.fullmatch(r"[0-9\.\+\-\*\/\(\)\s]+", expression):
        return float(fallback)

    try:
        return float(eval(expression, {"__builtins__": {}}, {}))
    except Exception:
        return float(fallback)


def split_top_level_args(arg_string: str) -> list[str]:
    args: list[str] = []
    current = []
    depth = 0
    for ch in arg_string:
        if ch == "(":
            depth += 1
        elif ch == ")":
            depth -= 1
        if ch == "," and depth == 0:
            args.append("".join(current).strip())
            current = []
            continue
        current.append(ch)
    if current:
        args.append("".join(current).strip())
    return args


def evaluate_excel_if_formula(
    formula: str | None,
    cell_values: dict[str, float],
    fallback: float,
) -> float:
    if not isinstance(formula, str):
        return float(fallback)

    expression = formula.strip()
    if expression.startswith("="):
        expression = expression[1:]
    expression = unwrap_iferror_expression(expression)

    if not expression.upper().startswith("IF(") or not expression.endswith(")"):
        return evaluate_excel_numeric_formula(formula, cell_values, fallback)

    inner = expression[3:-1]
    args = split_top_level_args(inner)
    if len(args) != 3:
        return float(fallback)

    condition_expr, true_expr, false_expr = args
    condition_match = re.match(r"^\s*([A-Za-z]+\d+)\s*([=<>]+)\s*([-]?[0-9]*\.?[0-9]+)\s*$", condition_expr)
    if not condition_match:
        return float(fallback)

    ref = condition_match.group(1).upper()
    op = condition_match.group(2)
    rhs = float(condition_match.group(3))
    lhs = float(cell_values.get(ref, 0.0) or 0.0)

    is_true = False
    if op == "=":
        is_true = lhs == rhs
    elif op == "<>":
        is_true = lhs != rhs
    elif op == ">":
        is_true = lhs > rhs
    elif op == "<":
        is_true = lhs < rhs
    elif op == ">=":
        is_true = lhs >= rhs
    elif op == "<=":
        is_true = lhs <= rhs

    selected_expr = true_expr if is_true else false_expr
    return evaluate_excel_numeric_formula(f"={selected_expr}", cell_values, fallback)


workbook_exists = WORKBOOK_PATH.exists()
json_exists = DATA_JSON_PATH.exists()

if workbook_exists:
    (
        intro_title,
        intro_text,
        disclaimer,
        defaults,
        components_df,
        source_df,
        source_row_groups,
        budget_to_source_blended_row,
        section_1_formulas,
        section_3_formulas,
    ) = load_budget_calculator_data_from_workbook(
        str(WORKBOOK_PATH),
        WORKBOOK_PATH.stat().st_mtime,
        str(DATA_JSON_PATH),
    )
elif json_exists:
    (
        intro_title,
        intro_text,
        disclaimer,
        defaults,
        components_df,
        source_df,
        source_row_groups,
        budget_to_source_blended_row,
        section_1_formulas,
        section_3_formulas,
    ) = load_budget_calculator_data_from_json(str(DATA_JSON_PATH), DATA_JSON_PATH.stat().st_mtime)
else:
    st.error("Neither JSON data file nor workbook file was found in the project root.")
    st.stop()

st.title("Rural Connectivity Budget Tool")
intro_text_display = str(intro_text).replace(
    "Segerberg & Noriega (2026)",
    "[Segerberg & Noriega (2026)](https://www.mdpi.com/2071-1050/18/6/2842)",
)
st.markdown(intro_text_display)
st.warning(disclaimer)

st.markdown("## Section 1: Transport Project Inputs")
left, right = st.columns([1, 1])

with left:
    st.text_input("Project Name / Description", value=str(defaults.get("project_name", "")))
    st.text_input("Country / Region", value=str(defaults.get("country_region", "")))
    total_budget_usd = st.number_input(
        "Total Transport Project Budget (USD)",
        min_value=0.0,
        value=float(defaults.get("total_budget_usd", 0.0)),
        step=1000000.0,
        format="%.0f",
    )

with right:
    allocation_pct_value = int(round(float(defaults["allocation_pct"]) * 100))
    allocation_pct_value = max(0, min(100, allocation_pct_value))
    allocation_pct_percent = st.slider(
        "% Allocated to RNTI / Rural Access",
        min_value=0,
        max_value=100,
        value=allocation_pct_value,
        step=1,
        help="Use whole-number percent (e.g., 7 for 7%).",
    )
    maintenance_pct_value = int(round(float(defaults["maintenance_pct"]) * 100))
    maintenance_pct_value = max(0, min(100, maintenance_pct_value))
    maintenance_pct_percent = st.slider(
        "Maintenance Reserve (%)",
        min_value=0,
        max_value=100,
        value=maintenance_pct_value,
        step=1,
        help="Recommended range in workbook: 5 to 15.",
    )

allocation_pct = allocation_pct_percent / 100
maintenance_pct = maintenance_pct_percent / 100

section_1_cells = {
    "C12": float(total_budget_usd),
    "C13": float(allocation_pct),
    "C16": float(maintenance_pct),
}
rnti_budget = evaluate_excel_numeric_formula(
    section_1_formulas.get("rnti_budget_available"),
    section_1_cells,
    fallback=float(total_budget_usd * allocation_pct),
)
section_1_cells["C14"] = float(rnti_budget)

maintenance_amount = evaluate_excel_numeric_formula(
    section_1_formulas.get("maintenance_reserve_amount"),
    section_1_cells,
    fallback=float(rnti_budget * maintenance_pct),
)
section_1_cells["C17"] = float(maintenance_amount)

capital_budget = evaluate_excel_numeric_formula(
    section_1_formulas.get("capital_budget_for_infrastructure"),
    section_1_cells,
    fallback=float(rnti_budget - maintenance_amount),
)

k1, k2, k3 = st.columns(3)
k1.metric("RNTI Budget Available", f"${rnti_budget:,.0f}")
k2.metric("Maintenance Reserve Amount", f"${maintenance_amount:,.0f}")
k3.metric("Capital Budget for Infrastructure", f"${capital_budget:,.0f}")

if "rnti_source_data_df" not in st.session_state:
    st.session_state["rnti_source_data_df"] = source_df.copy()
if "rnti_blended_overrides" not in st.session_state:
    st.session_state["rnti_blended_overrides"] = {}
if "rnti_qty_overrides" not in st.session_state:
    st.session_state["rnti_qty_overrides"] = {}
active_source_df = st.session_state["rnti_source_data_df"].copy()

blended_summary_df = build_blended_summary_table(active_source_df, source_row_groups)

blended_summary_df["Blended Min (USD)"] = to_numeric_series(blended_summary_df["Blended Min (USD)"])
blended_summary_df["Blended Max (USD)"] = to_numeric_series(blended_summary_df["Blended Max (USD)"])
blended_summary_df["Blended Mid (USD)"] = blended_summary_df.apply(
    lambda row: calculate_source_mid(row["Blended Min (USD)"], row["Blended Max (USD)"]),
    axis=1,
)

for budget_row, override_values in st.session_state["rnti_blended_overrides"].items():
    idx = blended_summary_df[blended_summary_df["Budget Row"] == int(budget_row)].index
    if idx.empty:
        continue
    override_low = to_numeric_series(pd.Series([override_values.get("Low Est. (USD)")])).iloc[0]
    override_high = to_numeric_series(pd.Series([override_values.get("High Est. (USD)")])).iloc[0]
    if not pd.isna(override_low):
        blended_summary_df.at[idx[0], "Blended Min (USD)"] = float(override_low)
    if not pd.isna(override_high):
        blended_summary_df.at[idx[0], "Blended Max (USD)"] = float(override_high)

blended_summary_df["Blended Mid (USD)"] = blended_summary_df.apply(
    lambda row: calculate_source_mid(row["Blended Min (USD)"], row["Blended Max (USD)"]),
    axis=1,
)

blended_override = {}
for _, row in blended_summary_df.iterrows():
    budget_row = int(row["Budget Row"])
    blended_override[budget_row] = (
        row["Blended Min (USD)"],
        row["Blended Max (USD)"],
        row["Blended Mid (USD)"],
    )

components_df = components_df.copy()
for idx in components_df.index:
    budget_row = int(components_df.at[idx, "Budget Row"])
    if budget_row in blended_override:
        min_cost, max_cost, mid_cost = blended_override[budget_row]
        components_df.at[idx, "Min Unit Cost (USD)"] = min_cost
        components_df.at[idx, "Max Unit Cost (USD)"] = max_cost
        components_df.at[idx, "Mid Unit Cost (USD)"] = mid_cost

st.markdown("## Section 2: Infrastructure Mix")

working_df = components_df.copy()

# Re-apply persisted Qty edits so multiple row edits are retained across reruns.
for budget_row, qty_value in st.session_state.get("rnti_qty_overrides", {}).items():
    row_idx = working_df[working_df["Budget Row"] == int(budget_row)].index
    if not row_idx.empty:
        working_df.at[row_idx[0], "Qty"] = qty_value

editor_state = st.session_state.get("rnti_infra_editor", {})
edited_rows = editor_state.get("edited_rows", {}) if isinstance(editor_state, dict) else {}


def section_editor_key(section_name: str) -> str:
    section_slug = re.sub(r"[^a-z0-9]+", "_", str(section_name).lower()).strip("_")
    return f"rnti_infra_editor_{section_slug}"


sections = [s for s in working_df["Section"].dropna().unique()]

if st.button("Reset all quantities to 0", key="rnti_reset_qty_btn"):
    st.session_state["rnti_qty_overrides"] = {
        int(budget_row): 0.0
        for budget_row in working_df["Budget Row"].dropna().tolist()
    }
    for section in sections:
        key = section_editor_key(str(section))
        if key in st.session_state and isinstance(st.session_state[key], dict):
            st.session_state[key]["edited_rows"] = {}
    st.rerun()

# Capture latest Qty edits from each section editor widget state.
for section in sections:
    key = section_editor_key(str(section))
    section_state = st.session_state.get(key, {})
    section_edited_rows = section_state.get("edited_rows", {}) if isinstance(section_state, dict) else {}

    section_rows = working_df[working_df["Section"] == section].reset_index(drop=True)
    row_index_to_budget_row = {
        idx: int(row["Budget Row"])
        for idx, (_, row) in enumerate(section_rows.iterrows())
    }

    for row_idx, changes in section_edited_rows.items():
        if not isinstance(changes, dict) or "Qty" not in changes:
            continue
        budget_row = row_index_to_budget_row.get(int(row_idx))
        if budget_row is None:
            continue
        qty_value = to_numeric_series(pd.Series([changes["Qty"]])).iloc[0]
        if pd.isna(qty_value):
            qty_value = 0.0
        st.session_state["rnti_qty_overrides"][int(budget_row)] = float(qty_value)

# Re-apply persisted + newly captured Qty edits.
for budget_row, qty_value in st.session_state.get("rnti_qty_overrides", {}).items():
    row_idx = working_df[working_df["Budget Row"] == int(budget_row)].index
    if not row_idx.empty:
        working_df.at[row_idx[0], "Qty"] = qty_value

working_df["Min Unit Cost (USD)"] = to_numeric_series(working_df["Min Unit Cost (USD)"])
working_df["Max Unit Cost (USD)"] = to_numeric_series(working_df["Max Unit Cost (USD)"])
working_df["Mid Unit Cost (USD)"] = to_numeric_series(working_df["Mid Unit Cost (USD)"])
working_df["Qty"] = to_numeric_series(working_df["Qty"]).fillna(0)
working_df["Total Cost Estimate (Mid)"] = working_df["Mid Unit Cost (USD)"] * working_df["Qty"]
working_df["Total Cost Estimate Range"] = working_df.apply(
    lambda row: calc_range_text(
        row["Min Unit Cost (USD)"],
        row["Max Unit Cost (USD)"],
        row["Qty"],
    ),
    axis=1,
)


def display_section_heading(section_name: str) -> str:
    mapping = {
        "ORDER 3 TERTIARY LINKS": "TERTIARY LINKS",
        "SUB-TERTIARY LINKS": "SUB-TERTIARY LINKS",
        "ANCILLARY & SAFETY INFRASTRUCTURE": "ANCILLARY INFRASTRUCTURE",
    }
    return mapping.get(section_name, str(section_name).upper())


st.caption("Edit values in the highlighted ✏️ Quantity column.")

edited_section_frames: list[pd.DataFrame] = []
for section in sections:
    st.markdown(f"#### {display_section_heading(str(section))}")
    section_editor_df = working_df[working_df["Section"] == section].copy()

    edited_section_df = st.data_editor(
        section_editor_df,
        hide_index=True,
        use_container_width=True,
        num_rows="fixed",
        key=section_editor_key(str(section)),
        column_order=[
            "Infrastructure Type",
            "Unit",
            "Qty",
            "Min Unit Cost (USD)",
            "Max Unit Cost (USD)",
            "Mid Unit Cost (USD)",
            "Total Cost Estimate (Mid)",
            "Total Cost Estimate Range",
        ],
        column_config={
            "Infrastructure Type": st.column_config.TextColumn(disabled=True, width="medium"),
            "Unit": st.column_config.TextColumn(disabled=True),
            "Mid Unit Cost (USD)": st.column_config.NumberColumn(disabled=True, format="$%0.0f"),
            "Qty": st.column_config.NumberColumn(
                "✏️ Quantity",
                min_value=0.0,
                step=1.0,
                format="%.2f",
                help="Primary editable input",
            ),
            "Min Unit Cost (USD)": st.column_config.NumberColumn(disabled=True, format="$%0.0f"),
            "Max Unit Cost (USD)": st.column_config.NumberColumn(disabled=True, format="$%0.0f"),
            "Total Cost Estimate (Mid)": st.column_config.NumberColumn(disabled=True, format="$%0.0f"),
            "Total Cost Estimate Range": st.column_config.TextColumn(disabled=True),
            "Budget Row": st.column_config.NumberColumn(disabled=True, format="%d"),
            "Section": st.column_config.TextColumn(disabled=True),
        },
    )

    edited_section_df["Min Unit Cost (USD)"] = to_numeric_series(edited_section_df["Min Unit Cost (USD)"])
    edited_section_df["Max Unit Cost (USD)"] = to_numeric_series(edited_section_df["Max Unit Cost (USD)"])
    edited_section_df["Mid Unit Cost (USD)"] = to_numeric_series(edited_section_df["Mid Unit Cost (USD)"])
    edited_section_df["Qty"] = to_numeric_series(edited_section_df["Qty"]).fillna(0)
    edited_section_df["Total Cost Estimate (Mid)"] = edited_section_df["Mid Unit Cost (USD)"] * edited_section_df["Qty"]
    edited_section_df["Total Cost Estimate Range"] = edited_section_df.apply(
        lambda row: calc_range_text(
            row["Min Unit Cost (USD)"],
            row["Max Unit Cost (USD)"],
            row["Qty"],
        ),
        axis=1,
    )

    edited_section_frames.append(edited_section_df)

editable_df = pd.concat(edited_section_frames, ignore_index=True) if edited_section_frames else working_df.copy()

summary_df = editable_df.copy()
summary_df["Min Unit Cost (USD)"] = to_numeric_series(summary_df["Min Unit Cost (USD)"])
summary_df["Max Unit Cost (USD)"] = to_numeric_series(summary_df["Max Unit Cost (USD)"])
summary_df["Mid Unit Cost (USD)"] = to_numeric_series(summary_df["Mid Unit Cost (USD)"])
summary_df["Qty"] = to_numeric_series(summary_df["Qty"]).fillna(0)
summary_df["Total Cost Estimate (Mid)"] = summary_df["Mid Unit Cost (USD)"] * summary_df["Qty"]
summary_df["Total Cost Estimate Range"] = summary_df.apply(
    lambda row: calc_range_text(
        row["Min Unit Cost (USD)"],
        row["Max Unit Cost (USD)"],
        row["Qty"],
    ),
    axis=1,
)

budget_row_mid_totals_source = summary_df.copy()
if "Budget Row" not in budget_row_mid_totals_source.columns:
    budget_row_mid_totals_source["Budget Row"] = pd.NA

budget_row_mid_totals_source["Budget Row"] = to_numeric_series(
    budget_row_mid_totals_source["Budget Row"]
)
if budget_row_mid_totals_source["Budget Row"].isna().all() and len(components_df) == len(budget_row_mid_totals_source):
    budget_row_mid_totals_source["Budget Row"] = components_df["Budget Row"].values

budget_row_mid_totals_source = budget_row_mid_totals_source[
    budget_row_mid_totals_source["Budget Row"].notna()
].copy()
budget_row_mid_totals_series = (
    budget_row_mid_totals_source.set_index("Budget Row")["Total Cost Estimate (Mid)"]
    .fillna(0)
)
budget_row_mid_totals = {
    int(row): float(value)
    for row, value in budget_row_mid_totals_series.items()
}

st.markdown("## Section 3: Budget Summary")
total_infrastructure_cost = calculate_total_infrastructure_cost_from_formula(
    section_3_formulas.get("total_infrastructure_cost"),
    budget_row_mid_totals,
)

section_3_cells = {
    "C12": float(total_budget_usd),
    "C14": float(rnti_budget),
    "C17": float(maintenance_amount),
    "C18": float(capital_budget),
    "C55": float(total_infrastructure_cost),
}

capital_budget_available = evaluate_excel_numeric_formula(
    section_3_formulas.get("capital_budget_available"),
    section_3_cells,
    fallback=float(capital_budget),
)
section_3_cells["C56"] = float(capital_budget_available)

remaining_capital_budget = evaluate_excel_numeric_formula(
    section_3_formulas.get("remaining_capital_budget"),
    section_3_cells,
    fallback=float(capital_budget_available - total_infrastructure_cost),
)
section_3_cells["C57"] = float(remaining_capital_budget)

pct_capital_used = evaluate_excel_if_formula(
    section_3_formulas.get("pct_capital_budget_used"),
    section_3_cells,
    fallback=0.0 if capital_budget == 0 else float(total_infrastructure_cost / capital_budget),
)
section_3_cells["C58"] = float(pct_capital_used)

maintenance_reserve = evaluate_excel_numeric_formula(
    section_3_formulas.get("maintenance_reserve"),
    section_3_cells,
    fallback=float(maintenance_amount),
)
section_3_cells["C59"] = float(maintenance_reserve)

total_estimated_rnti_investment = evaluate_excel_numeric_formula(
    section_3_formulas.get("total_estimated_rnti_investment"),
    section_3_cells,
    fallback=float(total_infrastructure_cost + maintenance_reserve),
)
section_3_cells["C60"] = float(total_estimated_rnti_investment)

pct_of_total_project = evaluate_excel_if_formula(
    section_3_formulas.get("pct_total_corridor_project"),
    section_3_cells,
    fallback=0.0 if total_budget_usd == 0 else float(total_estimated_rnti_investment / total_budget_usd),
)
section_3_cells["C61"] = float(pct_of_total_project)

s1, s2, s3, s4 = st.columns(4)
s1.metric("Total Infrastructure Cost (Mid-Point Estimate)", f"${total_infrastructure_cost:,.0f}")
s2.metric("Capital Budget Available", f"${capital_budget_available:,.0f}")
s3.metric("Remaining Capital Budget", f"${remaining_capital_budget:,.0f}")
s4.metric("% of Capital Budget Used", f"{pct_capital_used:.1%}")

s5, s6, s7 = st.columns(3)
s5.metric("Maintenance Reserve", f"${maintenance_reserve:,.0f}")
s6.metric("Total Estimated RNTI Investment", f"${total_estimated_rnti_investment:,.0f}")
s7.metric("% of Total Corridor Project", f"{pct_of_total_project:.4%}")

if remaining_capital_budget < 0:
    st.error(
        f"Your selected quantities exceed capital budget by ${abs(remaining_capital_budget):,.0f}."
    )
else:
    st.success("Selected quantities are within the current capital budget.")

st.markdown("## Quick Reference: Example Allocation Scenarios")
quick_ref = pd.DataFrame(
    {
        "% Allocation": ["1%", "2%", "5%", "10%"],
        "Budget = $100M": [1_000_000, 2_000_000, 5_000_000, 10_000_000],
        "Budget = $200M": [2_000_000, 4_000_000, 10_000_000, 20_000_000],
        "Budget = $400M": [4_000_000, 8_000_000, 20_000_000, 40_000_000],
        "Budget = $1B": [10_000_000, 20_000_000, 50_000_000, 100_000_000],
    }
)

st.dataframe(
    quick_ref,
    hide_index=True,
    use_container_width=True,
    column_config={
        "Budget = $100M": st.column_config.NumberColumn(format="$%0.0f"),
        "Budget = $200M": st.column_config.NumberColumn(format="$%0.0f"),
        "Budget = $400M": st.column_config.NumberColumn(format="$%0.0f"),
        "Budget = $1B": st.column_config.NumberColumn(format="$%0.0f"),
    },
)

if False:  # Keep Section 4 code for later, but hide it from the UI for now.
    st.markdown("### RNTI COST SOURCE DATA — Verifiable Sources with Integrated Blended Averages")

    grouped_source_frames = []
    new_blended_overrides = {}
    last_section = None
    for budget_row in source_group_display_order(source_row_groups):
        group_df = st.session_state["rnti_source_data_df"][
            st.session_state["rnti_source_data_df"]["Budget Row"] == budget_row
        ].copy()
        if group_df.empty:
            continue

        source_section = group_df["Source Section"].iloc[0]
        source_item_title = group_df["Source Item Title"].iloc[0]

        if source_section != last_section:
            st.markdown(f"#### {source_section}")
            last_section = source_section

        st.markdown(f"**{source_item_title}**")
        group_df["Source Mid (USD)"] = group_df.apply(
            lambda row: calculate_source_mid(row["Low Est. (USD)"], row["High Est. (USD)"]),
            axis=1,
        )

        blended_min = calculate_blended_average(group_df["Low Est. (USD)"])
        blended_max = calculate_blended_average(group_df["High Est. (USD)"])
        blended_mid = calculate_source_mid(blended_min, blended_max)

        edited_group_df = st.data_editor(
            group_df[
                [
                    "Geography / Context",
                    "Low Est. (USD)",
                    "High Est. (USD)",
                    "Source Mid (USD)",
                    "Source Name",
                    "Full Citation",
                    "Year",
                    "Budget Row",
                    "Source Row",
                    "Infrastructure Type",
                    "Source Section",
                    "Source Item Title",
                ]
            ],
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key=f"rnti_source_data_editor_{budget_row}",
            column_config={
                "Geography / Context": st.column_config.TextColumn(disabled=True, width="large"),
                "Low Est. (USD)": st.column_config.NumberColumn(format="$%0.0f"),
                "High Est. (USD)": st.column_config.NumberColumn(format="$%0.0f"),
                "Source Mid (USD)": st.column_config.NumberColumn(disabled=True, format="$%0.0f"),
                "Source Name": st.column_config.LinkColumn(disabled=True, width="medium"),
                "Full Citation": st.column_config.TextColumn(disabled=True, width="large"),
                "Year": st.column_config.NumberColumn(disabled=True, format="%d"),
                "Budget Row": st.column_config.NumberColumn(disabled=True, format="%d"),
                "Source Row": st.column_config.NumberColumn(disabled=True, format="%d"),
                "Infrastructure Type": st.column_config.TextColumn(disabled=True),
                "Source Section": st.column_config.TextColumn(disabled=True),
                "Source Item Title": st.column_config.TextColumn(disabled=True),
            },
            column_order=[
                "Geography / Context",
                "Low Est. (USD)",
                "High Est. (USD)",
                "Source Mid (USD)",
                "Source Name",
                "Full Citation",
                "Year",
            ],
        )

        edited_group_df = edited_group_df.reset_index(drop=True)

        edited_group_df["Low Est. (USD)"] = to_numeric_series(
            edited_group_df["Low Est. (USD)"]
        )
        edited_group_df["High Est. (USD)"] = to_numeric_series(
            edited_group_df["High Est. (USD)"]
        )
        edited_group_df["Source Mid (USD)"] = edited_group_df.apply(
            lambda row: calculate_source_mid(row["Low Est. (USD)"], row["High Est. (USD)"]),
            axis=1,
        )

        blended_low = calculate_blended_average(edited_group_df["Low Est. (USD)"])
        blended_high = calculate_blended_average(edited_group_df["High Est. (USD)"])
        blended_mid = calculate_source_mid(blended_low, blended_high)

        existing_override = st.session_state["rnti_blended_overrides"].get(int(budget_row), {})
        display_blended_low = to_numeric_series(pd.Series([existing_override.get("Low Est. (USD)")])).iloc[0]
        display_blended_high = to_numeric_series(pd.Series([existing_override.get("High Est. (USD)")])).iloc[0]
        if pd.isna(display_blended_low):
            display_blended_low = blended_low
        if pd.isna(display_blended_high):
            display_blended_high = blended_high
        display_blended_mid = calculate_source_mid(display_blended_low, display_blended_high)

        blended_editor_df = pd.DataFrame(
            [
                {
                    "Geography / Context": "BLENDED AVERAGE (Min / Max / Mid)",
                    "Low Est. (USD)": display_blended_low,
                    "High Est. (USD)": display_blended_high,
                    "Source Mid (USD)": display_blended_mid,
                    "Source Name": None,
                    "Full Citation": None,
                    "Year": None,
                }
            ]
        )

        edited_blended_df = st.data_editor(
            blended_editor_df,
            hide_index=True,
            use_container_width=True,
            num_rows="fixed",
            key=f"rnti_blended_editor_{budget_row}",
            column_order=[
                "Geography / Context",
                "Low Est. (USD)",
                "High Est. (USD)",
                "Source Mid (USD)",
                "Source Name",
                "Full Citation",
                "Year",
            ],
            column_config={
                "Geography / Context": st.column_config.TextColumn(disabled=True, width="large"),
                "Low Est. (USD)": st.column_config.NumberColumn(format="$%0.0f"),
                "High Est. (USD)": st.column_config.NumberColumn(format="$%0.0f"),
                "Source Mid (USD)": st.column_config.NumberColumn(disabled=True, format="$%0.0f"),
                "Source Name": st.column_config.LinkColumn(disabled=True, width="medium"),
                "Full Citation": st.column_config.TextColumn(disabled=True, width="large"),
                "Year": st.column_config.NumberColumn(disabled=True, format="%d"),
            },
        )

        edited_blended_low = to_numeric_series(pd.Series([edited_blended_df.at[0, "Low Est. (USD)"]])).iloc[0]
        edited_blended_high = to_numeric_series(pd.Series([edited_blended_df.at[0, "High Est. (USD)"]])).iloc[0]
        if pd.isna(edited_blended_low):
            edited_blended_low = blended_low
        if pd.isna(edited_blended_high):
            edited_blended_high = blended_high
        new_blended_overrides[int(budget_row)] = {
            "Low Est. (USD)": edited_blended_low,
            "High Est. (USD)": edited_blended_high,
        }

        grouped_source_frames.append(edited_group_df)

    editable_source_df = pd.concat(grouped_source_frames, ignore_index=True)
    editable_source_df["Low Est. (USD)"] = to_numeric_series(editable_source_df["Low Est. (USD)"])
    editable_source_df["High Est. (USD)"] = to_numeric_series(editable_source_df["High Est. (USD)"])
    editable_source_df["Source Mid (USD)"] = editable_source_df.apply(
        lambda row: calculate_source_mid(row["Low Est. (USD)"], row["High Est. (USD)"]),
        axis=1,
    )

    source_changed = not editable_source_df.equals(st.session_state["rnti_source_data_df"])
    overrides_changed = new_blended_overrides != st.session_state["rnti_blended_overrides"]

    if source_changed:
        st.session_state["rnti_source_data_df"] = editable_source_df.copy()
    if overrides_changed:
        st.session_state["rnti_blended_overrides"] = new_blended_overrides

    if source_changed or overrides_changed:
        st.rerun()

with st.expander("Methodology and caveats"):
    st.warning(
        """
- Unit costs are blended approximations from verifiable sources in the workbook.
- Items with no verifiable source are intentionally left blank.
- Actual costs vary by country, terrain, labor, materials, and design specification.
- This is a conversation/planning tool, not a final engineering cost estimate.
- Maintenance is represented as a reserve percentage, not full lifecycle costing.
"""
    )
